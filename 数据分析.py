import streamlit  streamlit  streamlit  streamlit as st st st st
import pandas  pandas  pandas  pandas as pd pd pd pd
import plotly.express  plotly.express  plotly.express  plotly.express as px px px px
import openpyxl openpyxl openpyxl openpyxl
from openpyxl.utils  openpyxl.utils  openpyxl.utils  openpyxl.utils import get_column_letter get_column_letter get_column_letter get_column_letter
import openai openai openai openai
import warnings warnings warnings warnings
warnings.warnings.warnings..warnings....warnings........filterwarnings('ignore')

# ---------------------- 配置 ChatGPT ----------------------
# 推荐用 Streamlit Secrets 管理 API Key（部署时在 Streamlit Cloud 配置）
openai.openai.openai..openai....api_key = st. = st. = st. = st. = st. = st. = st. = st.secrets........get("OPENAI_API_KEY", , , , , , , , "")

# 页面配置
st.st.st..set_page_config(
    page_title=    page_title="ChatGPT 增强版智能表格分析工具",,
    page_icon=    page_icon="🤖📊",,
    layout=    layout="wide"
)

st.st.st..st....st........title("🤖📊 ChatGPT 增强版智能表格分析工具")
st.st.st..st....st........markdown("### ✨ 任意格式表格 + 自然语言精准分析（支持复杂指令）")

# ---------------------- 核心1：智能表格解析（修复 openpyxl 版本兼容） ----------------------
def smart_parse_excel(file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=file, sheet_name=None)::::::::::::::::
    """智能解析Excel，自动定位有效数据，兼容任意格式"""
    if sheet_name  sheet_name  sheet_name  sheet_name is None::::
        xl_file = pd.                xl_file = pd.        xl_file = pd. = pd.ExcelFile(filefilefilefile)
        sheet_names = xl_file.        sheet_names = xl_file.sheet_names
        all_data =         all_data =         all_data =         all_data = {}
        for name  name  name  name in sheet_names: sheet_names: sheet_names: sheet_names:
            df =             df =             df =             df = parse_single_sheet(file, namefile, namefile, namefile, name)
            if not df.   df.  df. . df....empty::::::::
                all_data                all_data[namename] = df = df
        return all_data all_data all_data all_data
    else::::
        df =         df = parse_single_sheet(file, sheet_namefile, sheet_name)
        return {sheet_name: dfsheet_name: df}

def parse_single_sheet(file, sheet_namefile, sheet_name)::
    """解析单个sheet，处理合并单元格、空行空列（修复 openpyxl 版本兼容）"""
    wb = openpyxl.    wb = openpyxl.load_workbook(file, data_only=file, data_only=True)
    ws = wb    ws = wb[sheet_namesheet_name]
    
    # 修复：用 min_column/max_column 替代 min_col/max_col（新版 openpyxl 兼容）
    min_row, max_row = ws.    min_row, max_row = ws.min_row, ws., ws.max_row
    min_col, max_col = ws.    min_col, max_col = ws.min_column, ws., ws.max_column  # 关键修复行
    
    # 过滤全空行
    valid_rows =     valid_rows = []
    for row  row in range(min_row, max_row + min_row, max_row + 1)::
        row_data =         row_data = [ws.ws.cell(row=row, column=colrow=row, column=col)..value for col  col in range(min_col, max_col + min_col, max_col + 1)]
        if any(cell cell is not None and str(cellcell)..strip() !=  != "" for cell  cell in row_data row_data)::
            valid_rows.            valid_rows.append(rowrow)
    
    # 过滤全空列
    valid_cols =     valid_cols = []
    for col  col in range(min_col, max_col + min_col, max_col + 1)::
        col_data =         col_data = [ws.ws.cell(row=row, column=colrow=row, column=col)..value for row  row in valid_rows valid_rows]
        if any(cell cell is not None and str(cellcell)..strip() !=  != "" for cell  cell in col_data col_data)::
            valid_cols.            valid_cols.append(colcol)
    
    if not valid_rows  valid_rows or not valid_cols: valid_cols:
        return pd.DataFrame()
    
    # 提取表头和数据
    header_row = valid_rows[0]
    data_rows = valid_rows[1:]
    
    # 处理表头
    headers = []
    for col in valid_cols:
        cell = ws.cell(row=header_row, column=col)
        header = cell.value if cell.value is not None else f"列{get_column_letter(col)}"
        headers.append(str(header).strip())
    
    # 处理数据行（填充合并单元格）
    data = []
    for row in data_rows:
        row_vals = []
        for col in valid_cols:
            cell = ws.cell(row=row, column=col)
            if cell.coordinate in ws.merged_cells:
                for merged_range in ws.merged_cells:
                    if cell.coordinate in merged_range:
                        merged_cell = ws[merged_range.split(":")[0]]
                        row_vals.append(merged_cell.value)
                        break
            else:
                row_vals.append(cell.value)
        data.append(row_vals)
    
    # 构建并清洗DataFrame
    df = pd.DataFrame(data, columns=headers)
    df = df.replace("", None).dropna(how="all")
    # 自动转换数值列
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col], errors="ignore")
        except:
            pass
    
    return df

# ---------------------- 核心2：ChatGPT 自然语言解析 ----------------------
def chatgpt_parse_query(df, query):
    """调用 ChatGPT 解析自然语言指令，生成 Python 代码并执行分析（适配 OpenAI 1.0+ 新版本）"""
    # 生成精准提示词
    prompt = f"""
你是一个专业的数据分析助手，现在有一个 DataFrame，列名如下：{df.columns.tolist()}。
请根据用户的问题，生成可以直接在 Python 中执行的 Pandas 代码，仅输出代码，不要解释。
用户的问题是：{query}

注意：
- 数据框变量名为 df
- 只返回可执行的 Python 代码片段，不要包含任何解释或说明
- 如果需要可视化，使用 plotly.express，变量名为 fig
- 如果需要输出结果，将结果赋值给变量 result
- 确保代码可以直接运行，不要有语法错误
"""
    
    # 调用 OpenAI API（适配 1.0+ 新版本）
    try:
        # 初始化 OpenAI 客户端（新版必须用客户端方式调用）
        client = openai.OpenAI(api_key=st.secrets.get("OPENAI_API_KEY", ""))
        
        # 新版接口调用方式
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "你是一个专业的数据分析助手，擅长将自然语言转换为 Pandas 代码。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1  # 降低随机性，保证代码稳定性
        )
        # 新版获取返回内容的方式（注意是 .content 不是 ['content']）
        code = response.choices[0].message.content.strip()
        return code
    except Exception as e:
        st.error(f"调用 ChatGPT 出错：{str(e)}")
        return None

def execute_analysis(df, code):
    """执行 ChatGPT 生成的代码，返回分析结果"""
    local_vars = {"df": df, "px": px, "pd": pd}
    try:
        exec(code, globals(), local_vars)
        return local_vars.get("result", None), local_vars.get("fig", None)
    except Exception as e:
        st.error(f"代码执行出错：{str(e)}，请检查指令是否清晰")
        return None, None

# ---------------------- 主流程：上传文件 + 解析 + 交互 ----------------------
uploaded_files = st.file_uploader(
    "上传Excel文件（支持多个）",
    type=["xlsx", "xls"],
    accept_multiple_files=True
)

# 初始化会话状态
if "current_df" not in st.session_state:
    st.session_state.current_df = None

if uploaded_files:
    for file in uploaded_files:
        st.markdown(f"### 📄 解析文件：{file.name}")
        
        # 智能解析表格
        with st.spinner("正在智能解析表格数据..."):
            sheet_data_dict = smart_parse_excel(file)
        
        if not sheet_data_dict:
            st.warning(f"文件{file.name}未识别到有效数据，请检查表格内容")
            continue
        
        # 选择sheet
        sheet_names = list(sheet_data_dict.keys())
        selected_sheet = st.selectbox(
            f"选择{file.name}的sheet",
            sheet_names,
            key=f"sheet_{file.name}"
        )
        
        # 获取当前解析的DataFrame
        current_df = sheet_data_dict[selected_sheet]
        st.session_state.current_df = current_df
        
        # 1. 数据预览
        st.markdown(f"#### 📋 自动解析结果预览（{selected_sheet}）")
        st.dataframe(current_df, use_container_width=True)
        
        # 2. ChatGPT 自然语言分析对话框
        st.markdown(f"#### 🗣️ ChatGPT 自然语言精准分析")
        user_query = st.text_area(
            "请输入你的分析要求（支持复杂指令，比如：帮我找出华东区利润最高的3个产品，并计算它们的利润率）",
            placeholder="比如：计算各区域的平均利润并按从高到低排序 / 找出销售额超过100万的产品并展示占比",
            key=f"query_{file.name}",
            height=120
        )
        
        # 执行分析按钮
        if st.button(f"🤖 用 ChatGPT 分析", key=f"exec_{file.name}"):
            if not openai.api_key:
                st.error("请先在 Streamlit Secrets 中配置你的 OpenAI API Key！")
            elif user_query.strip() == "":
                st.warning("请输入分析要求后再执行！")
            else:
                with st.spinner("ChatGPT 正在思考并生成分析代码..."):
                    # 调用 ChatGPT 生成代码
                    code = chatgpt_parse_query(current_df, user_query)
                    if code:
                        st.markdown("#### 🧩 ChatGPT 生成的分析代码：")
                        st.code(code, language="python")
                        
                        # 执行代码
                        result, fig = execute_analysis(current_df, code)
                        
                        # 展示结果
                        st.markdown("#### 📊 分析结果：")
                        if result is not None:
                            st.dataframe(result, use_container_width=True)
                        if fig is not None:
                            st.plotly_chart(fig, use_container_width=True)
        
        # 3. 保留原有自动分析能力
        with st.expander("📈 点击展开：自动快速分析（无需输入指令）", expanded=False):
            numeric_cols = current_df.select_dtypes(include=["int64", "float64"]).columns.tolist()
            text_cols = current_df.select_dtypes(include=["object"]).columns.tolist()
            
            if numeric_cols:
                selected_num_col = st.selectbox("选择数值列快速分析", numeric_cols, key=f"auto_num_{file.name}")
                col1, col2 = st.columns(2)
                with col1:
                    fig_hist = px.histogram(current_df, x=selected_num_col, title=f"{selected_num_col}分布直方图")
                    st.plotly_chart(fig_hist, use_container_width=True)
                with col2:
                    fig_box = px.box(current_df, y=selected_num_col, title=f"{selected_num_col}箱线图（异常值）")
                    st.plotly_chart(fig_box, use_container_width=True)
            else:
                st.info("未识别到数值列，仅展示文本列频次统计")
                if text_cols:
                    selected_text_col = st.selectbox("选择文本列分析", text_cols, key=f"auto_text_{file.name}")
                    text_counts = current_df[selected_text_col].value_counts().head(20).reset_index()
                    text_counts.columns = [selected_text_col, "频次"]
                    fig = px.bar(text_counts, x=selected_text_col, y="频次", title=f"{selected_text_col}频次分布")
                    st.plotly_chart(fig, use_container_width=True)

else:
    st.info("✅ 请上传Excel文件（任意格式），支持：\n1. 自动解析无行列/格式限制\n2. ChatGPT 增强的自然语言精准分析")
import subprocess
import sys

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# 强制安装 openpyxl
try:
    import openpyxl
except ImportError:
    install("openpyxl==3.1.2")
    import openpyxl

